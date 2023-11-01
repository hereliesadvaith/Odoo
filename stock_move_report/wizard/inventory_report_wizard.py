# -*- coding: utf-8 -*
from odoo import fields, models
from odoo.exceptions import ValidationError, MissingError
import json
from odoo.tools import date_utils
import io
from openpyxl import Workbook, utils
from openpyxl.styles import Alignment, Font


class InventoryReportWizard(models.TransientModel):
    """
    Model for inventory report wizard
    """
    _name = "inventory.report.wizard"
    _description = "Inventory Report Wizard"

    product_ids = fields.Many2many("product.product", string="Products",
                                   help="Select Products")
    picking_type_ids = fields.Many2many("stock.picking.type", string="Transfer Type",
                                        help="Select Locations")
    state = fields.Selection(selection=[
        ("draft", "Draft"),
        ("waiting", "Waiting Another Operation"),
        ("confirmed", "Waiting"),
        ("assigned", "Ready"),
        ("done", "Done"),
        ("cancel", "Cancelled"),
    ],
        help="Status", string="Status")
    start_date = fields.Date(string="Start Date", help="Start Date")
    end_date = fields.Date(string="End Date", help="End Date")

    # Action Methods

    def print(self):
        """
        To print reports
        """
        company_id = self.env.context["allowed_company_ids"][0]
        if self.start_date and self.end_date:
            if self.start_date >= self.end_date:
                raise ValidationError("Set valid period")
        query = f"""
        SELECT stm.date, stm.reference, ptm.name AS product,
        stl.complete_name AS from, std.complete_name AS to,
        stm.product_uom_qty AS quantity, stm.state AS status,
        spt.name AS picking_type
        FROM stock_move AS stm
        INNER JOIN product_product AS ppr
        ON stm.product_id = ppr.id
        INNER JOIN product_template AS ptm
        ON ppr.product_tmpl_id = ptm.id
        INNER JOIN stock_location as stl
        ON stm.location_id = stl.id
        INNER JOIN stock_location as std
        ON stm.location_dest_id = std.id
        INNER JOIN stock_picking as spk
        ON stm.picking_id = spk.id
        INNER JOIN stock_picking_type as spt
        ON spk.picking_type_id = spt.id
        WHERE stm.company_id = {company_id}
        """
        if self.start_date:
            query += f"""
           and stm.date >= '{self.start_date}'
           """
        if self.end_date:
            query += f"""
            and stm.date <= '{self.end_date}'
            """
        if self.product_ids:
            products = tuple(self.product_ids.ids)
            if len(products) > 1:
                query += f"""
                and stm.product_id in {products}
                """
            else:
                query += f"""
                and stm.product_id = {self.product_ids.id}
                """
        if self.picking_type_ids:
            picking_types = tuple(self.picking_type_ids.ids)
            if len(picking_types) > 1:
                query += f"""
                and spk.picking_type_id in {picking_types}
                """
            else:
                query += f"""
                and spk.picking_type_id = {self.picking_type_ids.id}
                """
        if self.state:
            query += f"""
            and stm.state = '{self.state}'
            """
        query += "ORDER BY stm.date DESC"
        self.env.cr.execute(query)
        stock_moves = self.env.cr.dictfetchall()
        unique_picking_types = list(
            set([i["picking_type"]["en_US"] for i in stock_moves]))
        if len(stock_moves) == 0:
            raise MissingError("No records found")
        for move in stock_moves:
            move["status"] = move["status"].replace("_", " ").title()
        data = {
            "stock_moves": stock_moves,
            "unique_picking_types": unique_picking_types,
            "start_date": self.start_date,
            "end_date": self.end_date,
        }
        if self._context['button'] == "pdf":
            return self.env.ref("stock_move_report.inventory_report_action").report_action(
                None, data=data
            )
        elif self._context['button'] == "xlsx":
            return {
            "type": "ir.actions.report",
            "data": {
                "model": "inventory.report.wizard",
                "options": json.dumps(data,
                                      default=date_utils.json_default),
                "output_format": "xlsx",
                "report_name": "Stock Move Report",
            },
            "report_type": "xlsx",
            }

    def get_xlsx_report(self, data, response):
        """
        To generate xlsx report using openpyxl
        """
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        for column in range(1, 8):
            col_letter = utils.get_column_letter(column)
            sheet.column_dimensions[col_letter].width = 20
        sheet.column_dimensions['C'].width = 40
        head = Font(bold=True, size=20)
        thead = Font(bold=True)
        sheet.merge_cells('A1:G2')
        sheet['A1'] = "STOCK MOVE REPORT"
        sheet['A1'].font = head
        sheet['A1'].alignment = Alignment(horizontal='center')
        for cell in sheet[6]:
            cell.font = thead

        if data["start_date"]:
            sheet['A3'] = "Date From"
            sheet['B3'] = data["start_date"]
        if data["end_date"]:
            sheet['A4'] = "Date To"
            sheet['B4'] = data["end_date"]

        data_header = [
            ['Date', 'Reference', 'Product', 'Quantity', 'From', 'To', 'Status']
        ]
        for row_index, row in enumerate(data_header, start=6):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value).font = thead
        index = 7
        for type in data['unique_picking_types']:
            sheet.cell(row=index, column=1, value=type).font = thead
            index += 1
            for move in data['stock_moves']:
                if type == move['picking_type']['en_US']:
                    row = [
                        move['date'],
                        move['reference'],
                        move['product']['en_US'],
                        move['quantity'],
                        move['from'],
                        move['to'],
                        move['status'],
                    ]
                    for col, value in enumerate(row, start=1):
                        sheet.cell(row=index, column=col, value=value)
                    index += 1

        workbook.save(output)
        output.seek(0)
        response.stream.write(output.read())
        output.close()
